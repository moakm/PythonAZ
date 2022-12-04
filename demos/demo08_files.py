import os, re, csv, pandas
from lib.utils.util_console_mgmt import cls
from lib.utils.util_person import gender, birth_date


def demo_files():
    # 1. txt
    # 2. excel
    # 3. xml
    # 4. csv

    def txtfile():
        path = rf"{os.getcwd()}\files\data.txt"
        plik = open(path, 'r', encoding='utf-8')
        # print(plik.read())
        # print(plik.readline())
        # print(plik.readlines())
        # [print(w.strip("\n")) for w in plik.readlines()]

        # data_from_file = plik.readlines())
        # plik.close()
        # [print(w.strip("\n")) for w in data_from_file()]

        for w in plik:
            t = w.split(";")
            p = t[2].strip('\n')
            print(f"{t[0]} {t[1]} urodzon{'a' if gender(p) == 'W' else 'y'} {birth_date(p)}")
        plik.close()

        pattern = "(Neverm|Len)ore"

        # kody = ['windows-1250', 'utf-8', 'utf-16']
        # for k in kody:
        try:
            rows_list = []
            with open(rf"{os.getcwd()}\files\raven.txt", "r", encoding='utf-8') as f:
                for row in f:
                    t = row.strip().strip('\n').strip(" ")
                    if re.search(pattern, row): rows_list.append(t)

            for i, w in enumerate(rows_list, start=1):
                print(f"{i}. {w}")
        except:
            pass

        danebabcia = open(rf"{os.getcwd()}\files\danebabcia.txt", "r")
        d = danebabcia.read()
        pattern = "\d{4}"
        pin = re.search(pattern, d)
        if pin: print(pin.group())

        cls()

    def csvfile():
        # read
        with open(rf"{os.getcwd()}\files\data.csv", 'r') as f_csv:
            csv_read = csv.reader(f_csv, delimiter=";")

            for row in csv_read:
                print(row)

        with open(rf"{os.getcwd()}\files\data.csv", 'a') as f_csv:
            name = "Janusz"
            surname = "Nowak"
            pesel = "12121212345"
            write_row = csv.writer(f_csv, delimiter=';')
            write_row.writerow([name, surname, pesel])

    def xmlfile():

        """
        <xml deklaracja nagłówka>
        <auta>
            <auto rej="AB12345" vin="QWER123">
                <marka></marka>
                <model></model>
                <hp></hp>
                <cena waluta="EUR">10000</cena>
            </auto>
        </auta>
        :return:
        """

        from xml.etree.ElementTree import Element, SubElement, ElementTree, dump, parse
        def add_car(r, rej, vin, marka, model, moc, cena, waluta):
            auto = SubElement(r, 'auto', rej=rej, vin=vin)
            SubElement(auto, 'marka').text = marka
            SubElement(auto, 'model').text = model
            SubElement(auto, 'hp').text = f"{moc}"
            SubElement(auto, 'cena', waluta=waluta).text = f"{cena}"

        def delete_car(r, rej):
            for t in r.findall('auto'):
                rej_s = t.get('rej')
                if rej_s == rej:
                    r.remove(t)
                ElementTree(rt).write(rf"{os.getcwd()}\files\cars(1).xml", encoding='utf-8', xml_declaration=True)

        # create main element
        root = Element('auta')
        # add sub elements
        add_car(root, '077E', 'VW7433', 'VW', 'golf', 90, 3000, 'PLN')
        add_car(root, 'sr8064j', 'RG123986', 'Skoda', 'Octavia', 120, 10000, 'EUR')
        add_car(root, 'sr8065j', 'DS123876', 'VW', 'POLO', 200, 10000, 'PLN')
        add_car(root, 'SWD801a', 'PO12342131', 'Volvo', 'S80', 250, 60000, 'PLN')

        # zapis
        ElementTree(root).write(rf"{os.getcwd()}\files\cars.xml", encoding='utf-8', xml_declaration=True)

        # insert
        pl = parse(rf"{os.getcwd()}\files\cars.xml")
        rt = pl.getroot()
        add_car(rt, 'WS123971', 'SK1245a', 'Dacia', 'Duster', 80, 5000, 'PLN')
        ElementTree(rt).write(rf"{os.getcwd()}\files\cars.xml", encoding='utf-8', xml_declaration=True)

        # Read
        p = parse(rf"{os.getcwd()}\files\cars.xml")
        r = p.getroot()
        # 1. Iteration read
        # for t in r.iter():
        #     print(t.text, t.tag, t.attrib)

        # 2. Read by find -> xpath 'auto\marka'
        for t in r.findall('auto'):
            rej = t.get('rej')
            marka = t.find('marka').text
            model = t.find('model').text
            cena = t.find('cena')
            waluta = cena.get('waluta')
            wartosc = cena.text
            print(f"{marka}, {model} kosztuje {wartosc} {waluta}")

        # Delete
        delete_car(r, 'WS123971')

        cls()

    def xlsxfile():
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws1 = wb.create_sheet('Sheet1')
        ws2 = wb.create_sheet('mo4km', 0)
        ws.title = 'worksheet title'

        for sheet in wb:
            print(sheet.title)

        ws["A1"] = "imie"
        ws["B1"] = "nazwisko"
        ws["C1"] = "pesel"
        ws["A2"] = "Maria"
        ws["B2"] = "Warzecha"
        ws["C2"] = "12121212345"

        path = rf"{os.getcwd()}\files\data.txt"
        plik = open(path, 'r', encoding='utf-8')
        for i, w in enumerate(plik, start=3):
            t = w.split(";")
            #p = t[2].strip('\n')
            ws[f"A{i}"] = t[0]
            ws[f"B{i}"] = t[1]
            ws[f"C{i}"] = t[2]
        plik.close()

        wb.save(rf"{os.getcwd()}\files\excel_data.xlsx")
        wb.close()

        df = pandas.read_excel(rf"{os.getcwd()}\files\excel_data.xlsx", "worksheet title")
        print(df)

        cls()

    while True:
        print("1. TXT")
        print("2. csv")
        print("3. xml")
        print("4. excel")
        user_input = input("Chose an option: ")
        match user_input:
            case "1":
                txtfile()
            case "2":
                csvfile()
            case "3":
                xmlfile()
            case "4":
                xlsxfile()
            case _:
                break
